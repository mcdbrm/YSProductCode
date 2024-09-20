# import os
import sys
import datetime, time
from openpyxl import Workbook, load_workbook
from collections import defaultdict
# from openpyxl.utils import get_column_letter
# import argparse
import warnings
warnings.filterwarnings("ignore")

def input_and_exit():
    print("按下Enter回车键以退出程序...")
    input()
    sys.exit()

def find_duplicates(lst):
    element_count = defaultdict(int)
    for element in lst:
        element_count[tuple(element)] += 1

    duplicates = [[list(key), value] for key, value in element_count.items()]
    return duplicates

def open_workbook(file_path, sheet_names=None, data_only=True):
    print(f'正在打开-{file_path}')
    start_time = time.perf_counter()
    try:
        wb = load_workbook(file_path, data_only=data_only)
        end_time = time.perf_counter()  
        print(f'{file_path.split(".")[0]}.xlsx-打开时间:{end_time - start_time:.6f}秒')
        if sheet_names:
            ws_dict = {name: wb[name] for name in sheet_names}
            return wb, ws_dict
        else:
            return wb
    except Exception as e:
        print(e)
        print(f'{file_path}-打开失败')
        input_and_exit()

def find_brand_list(ws_cinfo):
    brand_list = []
    for i in range(2, ws_cinfo.max_row + 1):
        brand = ws_cinfo.cell(i, 2).value
        if brand is None:
            continue
        brand_list.append(brand)
    return list(set(brand_list))

# 根据某列的值查找并返回另一列的值
def lookup_value(ws, lookup_col, lookup_val, return_col):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, lookup_col).value == lookup_val:
            # print(ws.cell(row, return_col).value)
            return ws.cell(row, return_col).value
    return None

# 检查商品是否缺少必需的字段，若缺少则退出程序
def validate_commodity_data(commodity, required_keys):
    for key in required_keys:
        if commodity.get(key) is None:
            print(f"商品数据缺失: {key} 查询失败")
            input_and_exit()

# 处理单件装的商品信息
def process_single_item(ws_sinfo, commodity):
    for row in range(2, ws_sinfo.max_row + 1):
        if ws_sinfo.cell(row, 4).value == commodity.get('品类'):
            commodity['商品分类'] = ws_sinfo.cell(row, 3).value
            commodity['季节'] = ws_sinfo.cell(row, 1).value
            break
    validate_commodity_data(commodity, ['商品分类', '季节'])

# 处理多件装的商品信息
def process_multi_item(ws_sinfo, ws_minfo, commodity):
    for row in range(2, ws_sinfo.max_row + 1):
        if ws_sinfo.cell(row, 4).value == commodity.get('品类'):
            commodity['商品分类'] = ws_sinfo.cell(row, 3).value
            commodity['季节'] = ws_sinfo.cell(row, 1).value
            commodity['单件组合装款式编码'] = ws_sinfo.cell(row, 6).value
            break
    validate_commodity_data(commodity, ['商品分类', '季节', '单件组合装款式编码'])

    commodity['组合装款式商品编码'] = lookup_value(ws_minfo, 4, commodity.get('组合装款式编码'), 3)
    validate_commodity_data(commodity, ['组合装款式商品编码'])

# 处理印花相关信息，side 为 '前' 或 '后'
def process_print_data(ws_pinfo, commodity, side):
    position_key = f'位置-{side}'
    print_name_key = f'印花名称-{side}'
    print_code_key = f'印花编码-{side}'
    position_code_key = f'位置代码-{side}'

    if commodity.get(print_name_key) is not None:
        if commodity.get(position_key) in ["胸", "裤"]:
            commodity[print_code_key] = commodity.get(print_name_key) + 'X'
            position_col, code_col = 10, 11
        else:
            commodity[print_code_key] = commodity.get(print_name_key) + 'D'
            position_col, code_col = 7, 8

        # print(f"\n[DEBUG] side: {side}")
        # print(f"[DEBUG] {position_key}: {commodity.get(position_key)}")
        # print(f"[DEBUG] {print_name_key}: {commodity.get(print_name_key)}")
        # print(f"[DEBUG] {print_code_key}: {commodity[print_code_key]}")
        # print(f"[DEBUG] position_col: {position_col}, code_col: {code_col}")
        commodity[position_code_key] = lookup_value(ws_pinfo, position_col, commodity.get(position_key), code_col)
        validate_commodity_data(commodity, [position_code_key])

        commodity[print_name_key] = lookup_value(ws_pinfo, 4, commodity.get(print_code_key), 1)
        validate_commodity_data(commodity, [print_name_key])

# 处理不同品牌的商品数据（除HL之外）
def process_brand_data(commodity, workbook, sheet_name, specification_brand, brand_code, range_columns, gender=None):
    if commodity.get('组合形式') == '多件装':
        temp = commodity.get('组合装款式商品编码')
        commodity['组合装款式商品编码'] = temp[0:4] + brand_code + temp[-1:]

    try:
        worksheet = workbook[sheet_name]
    except Exception as e:
        print(e)
        print(f"{brand_code} 吊牌信息-打开失败: {sheet_name} 表")
        input_and_exit()

    trademark_list = []
    for li in range(2, worksheet.max_row + 1):
        if worksheet.cell(li, 1).value == commodity.get('单件组合装款式编码') and worksheet.cell(li, 2).value == specification_brand:
            trademark_list = [worksheet.cell(li, j).value for j in range(range_columns[0], range_columns[1])]
    
    if len(trademark_list) == 0:
        print(f"{brand_code} 吊牌信息汇总表查询失败-《{commodity.get('品类')}》分表中找不到{commodity.get('单件组合装款式编码')} {specification_brand}")
        input_and_exit()
    else:
        return trademark_list

# 添加单件组合装信息到组合表中
def append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, code, trademark_list, ws_single_combination, s):
    list_single_combination = [
        commodity.get('单件组合装款式编码'),
        combination_commodity_code,
        s_commodity_name,
        entity_code,
        '成品',
        specification,
        code,
        1,
        0,
        'YS'
    ] + trademark_list

    if combination_commodity_code + code not in s:
        s.add(combination_commodity_code + code)
        ws_single_combination.append(list_single_combination)

# 临时编码
def generate_temp_codes(commodity, position_prefix, color):
    if commodity.get(f'位置-{position_prefix}') == '大图':
        commodity[f'位置-{position_prefix}'] = ''
    temp_code1 = commodity.get(f'印花名称-{position_prefix}') + commodity.get(f'位置代码-{position_prefix}') + '/' + color
    temp_code2 = commodity.get(f'印花名称-{position_prefix}') + commodity.get(f'位置-{position_prefix}') + '/' + color
    return temp_code1, temp_code2

# 临时编码追加到对应列表
def append_codes(picture_position_color_list, picture_color_list, color_list, temp_code1, temp_code2, color):
    picture_position_color_list.append(temp_code1)
    picture_color_list.append(temp_code2)
    color_list.append(color)

if __name__ == "__main__":
    # parser = argparse.ArgumentParser(description='YS公司自动化程序')
    # parser.add_argument('file1', help='xx文件')
    # parser.add_argument('file2', type=int, help='xx文件')
    # args = parser.parse_args()

    # path = os.path.dirname(os.path.abspath(__file__))

    print('*************************************')
    print('*       YS商品编码器程序v1.0        *')
    print('*************************************')

    wb_cinfo, ws_cinfo_dict = open_workbook('商品编码信息表.xlsx', ['商品编码信息表1'])
    ws_cinfo = ws_cinfo_dict['商品编码信息表1']

    wb_baseinfo, ws_baseinfo_dict = open_workbook('资料生成器.xlsx', ['附表1印花基础资料', '附表2单品基础资料', '附表3多件装基础信息', '大货称重表'])
    ws_pinfo = ws_baseinfo_dict['附表1印花基础资料']
    ws_sinfo = ws_baseinfo_dict['附表2单品基础资料']
    ws_minfo = ws_baseinfo_dict['附表3多件装基础信息']
    ws_winfo = ws_baseinfo_dict['大货称重表']

    wb_relationship, wb_relationship_dict = open_workbook('商品对应关系.xlsx', ['Sheet1'])
    ws_relationship = wb_relationship['Sheet1']
    
    brand_list = find_brand_list(ws_cinfo)
    if brand_list:
        if 'HL' in brand_list:
            wb_HL, ws_HL_type_dict = open_workbook('回力吊牌信息汇总表.xlsx', ['回力童装号型对照表'])
            ws_HL_type = ws_HL_type_dict['回力童装号型对照表']

        if 'BD' in brand_list:
            wb_BD = open_workbook('巴帝吊牌信息汇总表.xlsx')

        if 'SE' in brand_list:
            wb_SE = open_workbook('少宜吊牌信息汇总表.xlsx')
        
        if 'ML' in brand_list:
            wb_ML_male = open_workbook('菲尔吊牌信息汇总表-男童.xlsx')
            wb_ML_female = open_workbook('菲尔吊牌信息汇总表-女童.xlsx')

        if 'JW' in brand_list:
            wb_JW_male = open_workbook('真维斯吊牌信息汇总表-男童.xlsx')
            wb_JW_female = open_workbook('真维斯吊牌信息汇总表-女童.xlsx')

    # 创建汇总表
    wb_final = Workbook()
    ws_single_general = wb_final.active
    ws_single_general.title = '单品-普通资料'
    ws_single_general.append(['款式编码','商品编码','商品名','分类','颜色及规格', '重量', '品牌', '虚拟分类', '国标码', \
                              '其它属性1', '其它属性2', '其它属性3', '其它属性4', '其它属性5', '其它属性6', '其它属性7', \
                                '其它属性8', '其它属性9', '其它属性10'])
    ws_single_combination = wb_final.create_sheet('单品-组合构成')
    ws_single_combination.append(['组合款式编码', '组合商品编码', '组合商品名称', '组合商品实体编码', '虚拟分类', \
                                  '组合颜色规格',  '商品编码', '数量', '应占售价', '品牌', '组合装国标码', \
                                    '其它属性1', '其它属性2', '其它属性3', '其它属性4', '其它属性5', '其它属性6', \
                                        '其它属性7', '其它属性8', '其它属性9', '其它属性10'])
    ws_multiple_combination = wb_final.create_sheet('多件装-组合构成')
    ws_multiple_combination.append(['组合款式编码', '组合商品编码', '组合商品名称', '虚拟分类', '组合颜色规格', \
                                    '商品编码', '数量', '应占售价', '品牌'])
    s = set()
    ws_multiple_combination_tmplst = []

    # 设置列宽
    # lst = [ws_single_general, ws_single_combination, ws_multiple_combination]
    # for l in lst:
    #     for i in range(1, l.max_column + 1):
    #         l.column_dimensions[get_column_letter(i)].width = 20.0
    

    # 关系对照表从第二行开始填充数据
    relationship_row = 2
    commodities = []

    for ci in range(4, ws_cinfo.max_row + 1): 
        print(f'_________________读第{ci}行________________')
        if  ws_cinfo.cell(ci, 2).value == None:
            continue

        count = 0
        # 遍历G到J列
        for i in range(7, 11):
            if ws_cinfo.cell(ci, i).value is None or ws_cinfo.cell(ci, i).value is 0:
                continue
            item = {}
            item['品类'] = ws_cinfo.cell(ci, i).value
            item['颜色'] = ws_cinfo.cell(ci, 5+i+count*4).value
            tmp = ws_cinfo.cell(ci, 6+i+count*4).value
            item['印花名称-前'] = str(tmp) if tmp is not None else tmp
            item['位置-前'] = ws_cinfo.cell(ci, 7+i+count*4).value
            tmp = ws_cinfo.cell(ci, 8+i+count*4).value
            item['印花名称-后'] = str(tmp) if tmp is not None else tmp

            item['位置-后'] = ws_cinfo.cell(ci, 9+i+count*4).value
            
            item['单件组合装款式编码'] = ws_cinfo.cell(ci, 5).value
            item['组合装款式编码'] = ws_cinfo.cell(ci, 6).value
            item['品牌'] = ws_cinfo.cell(ci, 2).value
            item['组合形式'] = ws_cinfo.cell(ci, 3).value
            item['性别'] = ws_cinfo.cell(ci, 4).value
            item['尺码'] = ws_cinfo.cell(ci, 11).value
            item['数量'] = ''
            
            count = count + 1
            commodities.append(item)

        # print(commodities)
        # sys.exit()
        for commodity in commodities:
            print(commodity)
            # 处理单件装和多件装
            if len(commodities) == 1 and commodity.get('组合形式') == '单件装':
                process_single_item(ws_sinfo, commodity)
            elif commodity.get('组合形式') == '多件装':
                process_multi_item(ws_sinfo, ws_minfo, commodity)

            # 处理前后的印花信息
            process_print_data(ws_pinfo, commodity, '前')
            process_print_data(ws_pinfo, commodity, '后')

            # print(commodities)
            # sys.exit()
            for size in commodity.get('尺码').split('/'):
                # 如果只有前印花编码
                if commodity.get('印花编码-前') and not commodity.get('印花编码-后'):
                    position_front = '' if commodity.get('位置-前') == '大图' else commodity.get('位置-前')
                    commodity_code = f"{commodity.get('印花编码-前')}{position_front}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    s_commodity_name = f"{commodity.get('印花名称-前')}{position_front}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    combination_commodity_code = f"{commodity.get('印花编码-前')}{position_front}/{commodity.get('品牌')}{commodity.get('品类')}/{commodity.get('颜色')}/{size}"
                    entity_code = commodity_code
                    # 回力查询关键字
                    trademark_HL_A1 = f"{commodity.get('印花编码-前')}{position_front}/{commodity.get('品类')}/{commodity.get('颜色')}"

                # 如果只有后印花编码
                elif not commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    position_back = '' if commodity.get('位置-后') == '大图' else commodity.get('位置-后')

                    commodity_code = f"{commodity.get('印花编码-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    s_commodity_name = f"{commodity.get('印花名称-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    combination_commodity_code = f"{commodity.get('印花编码-后')}{position_back}/{commodity.get('品牌')}{commodity.get('品类')}/{commodity.get('颜色')}/{size}"
                    entity_code = commodity_code
                    # 回力查询关键字
                    trademark_HL_A1 = f"{commodity.get('印花编码-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}"

                # 如果前后印花编码都有
                elif commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    position_front = '' if commodity.get('位置-前') == '大图' else commodity.get('位置-前')
                    position_back = '' if commodity.get('位置-后') == '大图' else commodity.get('位置-后')

                    commodity_code = f"{commodity.get('印花编码-前')}{position_front}_{commodity.get('印花编码-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    s_commodity_name = f"{commodity.get('印花名称-前')}{position_front}_{commodity.get('印花名称-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    combination_commodity_code = f"{commodity.get('印花编码-前')}{position_front}_{commodity.get('印花编码-后')}{position_back}/{commodity.get('品牌')}{commodity.get('品类')}/{commodity.get('颜色')}/{size}"
                    entity_code = commodity_code
                    # 回力查询关键字
                    trademark_HL_A1 = f"{commodity.get('印花编码-前')}{position_front}_{commodity.get('印花编码-后')}{position_back}/{commodity.get('品类')}/{commodity.get('颜色')}"
                
                if commodity.get('品牌') == 'CP':
                    combination_commodity_code = combination_commodity_code.replace('CP', '')

                specification = f"{commodity.get('颜色')};{size}"
                specification_brand = f"{commodity.get('颜色')};{size}-{commodity.get('品牌')}"
                single_combination_code = f"纯色/{commodity.get('品类')}/{commodity.get('颜色')}/{size}"

                # 称重表查询
                for wi in range(2, ws_winfo.max_row + 1):
                    if ws_winfo.cell(wi, 1).value == commodity.get('品类') and str(ws_winfo.cell(wi, 2).value) == str(size):
                        commodity['重量'] = ws_winfo.cell(wi, 3).value
                        break
                if commodity.get('重量') is None:
                    print(f"称重表查询失败: {commodity.get('品类')} {size}")
                    input_and_exit()

                # 各品牌处理
                trademark_list = []

                if commodity.get('品牌') == 'HL':
                    # print(f'trademark_HL_A1:{trademark_HL_A1}')
                    if commodity.get('组合形式') == '多件装':
                        commodity['组合装款式商品编码'] = commodity.get('组合装款式商品编码')[0:4] + 'HL' + commodity.get('组合装款式商品编码')[-1:]

                    # 不拆表
                    try:
                        ws_HL = wb_HL[commodity.get('品类')]
                    except Exception as e:
                        print(e)
                        print(f"回力吊牌信息-打开失败:{commodity.get('品类')}表")
                        print("按下Enter回车键以退出程序...")
                        input()
                        sys.exit()

                    HL_type = ''
                    for li in range(2, ws_HL_type.max_row + 1):
                        # print(f'{ws_HL_type.cell(li, 1).value}/{ws_HL_type.cell(li, 2).value}')
                        if ws_HL_type.cell(li, 1).value == commodity.get('品类') and ws_HL_type.cell(li, 2).value == int(size):
                            HL_type = ws_HL_type.cell(li, 3).value
                    if len(HL_type) == 0:
                        print(f"回力童装号型对照表查询失败:{commodity.get('品类')}/{size}")
                        print("按下Enter回车键以退出程序...")
                        input()
                        sys.exit()

                    for bi in range(2, ws_HL.max_row + 1):
                        if ws_HL.cell(bi, 1).value == trademark_HL_A1:
                            trademark_list = [ws_HL.cell(bi, j).value for j in range(2, 10)]
                            trademark_list.insert(1, HL_type)
                            trademark_list.insert(6, size)
                    if len(trademark_list) == 0:
                        print('回力吊牌信息汇总表查询失败')
                        print("按下Enter回车键以退出程序...")
                        input()
                        sys.exit()

                if commodity.get('品牌') == 'BD':
                    trademark_list = process_brand_data(commodity, wb_BD, commodity.get('品类'), specification_brand, 'BD', (3, 12))

                if commodity.get('品牌') == 'SE':
                    trademark_list = process_brand_data(commodity, wb_SE, commodity.get('品类'), specification_brand, 'SE', (3, 12))

                if commodity.get('品牌') == 'ML' and commodity.get('性别') == '男':
                    trademark_list = process_brand_data(commodity, wb_ML_male, commodity.get('品类'), specification_brand, 'ML', (3, 14))

                if commodity.get('品牌') == 'ML' and commodity.get('性别') == '女':
                    trademark_list = process_brand_data(commodity, wb_ML_female, commodity.get('品类'), specification_brand, 'ML', (3, 14))

                if commodity.get('品牌') == 'JW' and commodity.get('性别') == '男':
                    trademark_list = process_brand_data(commodity, wb_JW_male, commodity.get('品类'), specification_brand, 'JW', (3, 14))

                if commodity.get('品牌') == 'JW' and commodity.get('性别') == '女':
                    trademark_list = process_brand_data(commodity, wb_JW_female, commodity.get('品类'), specification_brand, 'JW', (3, 14))
                    
                
                # 通用部分处理 （款式编码 商品编码 商品名 分类 颜色及规格 重量 品牌 虚拟分类）
                list_single_general = [
                    commodity.get('单件组合装款式编码'),
                    commodity_code,
                    s_commodity_name,
                    commodity.get('商品分类'),
                    specification_brand,
                    commodity.get('重量'),
                    'YS',
                    commodity.get('季节')
                ] + trademark_list

                if commodity_code not in s:
                    s.add(commodity_code)
                    ws_single_general.append(list_single_general)

                # 单件装处理 
                if len(commodities) == 1 and commodity.get('组合形式') == '单件装':
                    ws_relationship.cell(relationship_row, 8).value = commodity_code
                    relationship_row += 1

                # 处理印花编码-前/后情况  （组合款式编码 组合商品编码 组合商品名称 组合商品实体编码 虚拟分类 组合颜色规格 商品编码 数量 应占售价 品牌）
                if commodity.get('印花编码-前') and not commodity.get('印花编码-后'):
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, commodity.get('印花编码-前'), trademark_list, ws_single_combination, s)
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, single_combination_code, trademark_list, ws_single_combination, s)

                elif not commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, commodity.get('印花编码-后'), trademark_list, ws_single_combination, s)
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, single_combination_code, trademark_list, ws_single_combination, s)

                elif commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, commodity.get('印花编码-前'), trademark_list, ws_single_combination, s)
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, commodity.get('印花编码-后'), trademark_list, ws_single_combination, s)
                    append_combination(commodity, combination_commodity_code, s_commodity_name, entity_code, specification, single_combination_code, trademark_list, ws_single_combination, s)

        if commodity.get('组合形式') == '多件装':
            picture_position_color_list = []
            picture_color_list = []
            color_list = []

            # 处理每个商品的印花编码及颜色
            for commodity in commodities:
                color = commodity.get('颜色')
                
                # 仅有前印花
                if commodity.get('印花编码-前') and not commodity.get('印花编码-后'):
                    temp_code1, temp_code2 = generate_temp_codes(commodity, '前', color)
                    append_codes(picture_position_color_list, picture_color_list, color_list, temp_code1, temp_code2, color)

                # 仅有后印花
                elif not commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    temp_code1, temp_code2 = generate_temp_codes(commodity, '后', color)
                    append_codes(picture_position_color_list, picture_color_list, color_list, temp_code1, temp_code2, color)

                # 前后印花都有
                elif commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                    temp_code1 = f"{commodity.get('印花名称-前')}{commodity.get('位置代码-前')}_{commodity.get('印花名称-后')}{commodity.get('位置代码-后')}/{color}"
                    temp_code2 = f"{commodity.get('印花名称-前')}{commodity.get('位置-前')}_{commodity.get('印花名称-后')}{commodity.get('位置-后')}/{color}"
                    append_codes(picture_position_color_list, picture_color_list, color_list, temp_code1, temp_code2, color)

            picture_position_color = '-'.join(picture_position_color_list)
            picture_color = '-'.join(picture_color_list)
            combination_color = '-'.join(color_list)

            # 补充关系对应表
            for size in commodity.get('尺码').split('/'):
                combination_commodity_code = f"{commodity.get('组合装款式商品编码')}-{picture_position_color}-{size}"
                ws_relationship.cell(relationship_row, 8).value = combination_commodity_code
                relationship_row += 1

            # 处理每个商品的组合装信息
            for commodity in commodities:
                for size in commodity.get('尺码').split('/'):
                    combination_commodity_code = f"{commodity.get('组合装款式商品编码')}-{picture_position_color}-{size}"
                    combination_commodity_name = f"{commodity.get('组合装款式商品编码')}-{picture_color}-{size}"
                    combination_color_code = f"{combination_color};{size}"

                    # 生成商品编码
                    if commodity.get('印花编码-前') and not commodity.get('印花编码-后'):
                        commodity_code = f"{commodity.get('印花编码-前')}{commodity.get('位置-前')}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    elif not commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                        commodity_code = f"{commodity.get('印花编码-后')}{commodity.get('位置-后')}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"
                    elif commodity.get('印花编码-前') and commodity.get('印花编码-后'):
                        commodity_code = f"{commodity.get('印花编码-前')}{commodity.get('位置-前')}_{commodity.get('印花编码-后')}{commodity.get('位置-后')}/{commodity.get('品类')}/{commodity.get('颜色')}/{size}-{commodity.get('品牌')}"

                    # 组合款式编码 组合商品编码 组合商品名称 虚拟分类 组合颜色规格 商品编码 数量 应占售价 品牌
                    list_multiple_combination = [
                        commodity.get('组合装款式编码'),
                        combination_commodity_code,
                        combination_commodity_name,
                        '成品',
                        combination_color_code,
                        commodity_code,
                        1,
                        0,
                        'YS'
                    ]
                    ws_multiple_combination_tmplst.append(list_multiple_combination)

    if ws_multiple_combination_tmplst:
        duplicates = find_duplicates(ws_multiple_combination_tmplst)
        list_multiple_combination.clear()
        for duplicate in duplicates:
                # print(duplicate)
                list_multiple_combination = duplicate[0]
                list_multiple_combination[-3] = duplicate[1]
                ws_multiple_combination.append(list_multiple_combination)

    wb_cinfo.close()
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    print(' ')
    print('_________________制作完成________________')
    print(f'编码生成表_{now}.xlsx')
    wb_final.save(f'编码生成表_{now}.xlsx')

    print(f'关系对应生成表_{now}.xlsx')
    wb_relationship.save(f'关系对应生成表_{now}.xlsx')

    print(" ")
    print("按下Enter回车键以退出程序...")
    input()